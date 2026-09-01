import io
from datetime import datetime
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.dashboard_service import dashboard_service


def get_val(obj, key, default=None):
    """Helper para obtener valor ya sea de dict o de dataclass/objeto."""
    if obj is None:
        return default
    if isinstance(obj, dict):
        return obj.get(key, default)
    return getattr(obj, key, default)


class ReportService:
    def generate_pdf_report(self, db: Session, user_id: int) -> io.BytesIO:
        """Genera un reporte PDF completo para el usuario dado."""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36
        )

        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=20,
            leading=24,
            textColor=colors.HexColor('#1e1b4b'),
            spaceAfter=6
        )
        subtitle_style = ParagraphStyle(
            'ReportSubtitle',
            parent=styles['Normal'],
            fontSize=10,
            leading=12,
            textColor=colors.HexColor('#64748b'),
            spaceAfter=15
        )
        section_heading = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=12,
            leading=16,
            textColor=colors.HexColor('#312e81'),
            spaceBefore=12,
            spaceAfter=6
        )

        # Datos del dashboard
        summary = dashboard_service.get_user_summary(db, user_id)
        auth_metrics = dashboard_service.get_auth_metrics(db, user_id)
        models = dashboard_service.get_model_versions(db, user_id)
        adapt_info = dashboard_service.get_adaptation_summary(db, user_id)
        timeline = dashboard_service.get_adaptation_timeline(db, user_id)
        comparison = dashboard_service.get_comparison_static_vs_adaptive(db, user_id)

        username = get_val(summary, 'username', f'User_{user_id}')

        # Encabezado
        elements.append(Paragraph("TECLEOLLAVE-ADAPT — Reporte Biométrico", title_style))
        elements.append(Paragraph(
            f"Usuario: <b>{username}</b> (ID: {user_id}) | Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            subtitle_style
        ))

        # 1. Resumen de KPIs
        elements.append(Paragraph("1. Métricas Clave y KPIs", section_heading))
        
        far_val = get_val(auth_metrics, 'far', 0)
        frr_val = get_val(auth_metrics, 'frr', 0)
        avg_score_val = get_val(auth_metrics, 'avg_score', 0)
        allow_c = get_val(auth_metrics, 'allow_count', 0)
        reject_c = get_val(auth_metrics, 'reject_count', 0)
        total_att = get_val(auth_metrics, 'total_attempts', 0)

        kpi_data = [
            ["Métrica", "Valor", "Métrica", "Valor"],
            ["Modelo Activo", f"v{get_val(summary, 'active_model_version', 'N/A')}", "FAR (30d)", f"{(far_val*100):.2f}%"],
            ["Total Muestras", str(get_val(summary, 'total_samples', 0)), "FRR (30d)", f"{(frr_val*100):.2f}%"],
            ["Intentos Auth", str(total_att), "Score Promedio", f"{avg_score_val:.3f}"],
            ["Adaptaciones", str(get_val(summary, 'total_adaptations', 0)), "Permitidos / Rechazados", f"{allow_c} / {reject_c}"]
        ]
        t_kpi = Table(kpi_data, colWidths=[130, 130, 130, 130])
        t_kpi.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#e0e7ff')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#1e1b4b')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(t_kpi)
        elements.append(Spacer(1, 10))

        # 2. Historial de Modelos
        elements.append(Paragraph("2. Historial de Versiones de Modelo", section_heading))
        model_rows = [["Versión", "Estado", "Muestras", "Intentos", "Tasa Permitidos", "Score Prom.", "Creado"]]
        for m in models or []:
            v_id = get_val(m, 'version_id')
            is_act = get_val(m, 'is_active')
            tr_samples = get_val(m, 'training_samples', 0)
            a_cnt = get_val(m, 'auth_count', 0)
            a_rate = get_val(m, 'allow_rate', 0)
            a_score = get_val(m, 'avg_score', 0)
            created = str(get_val(m, 'created_at', '')).split('T')[0]

            status_text = "ACTIVO" if is_act else "Archivado"
            model_rows.append([
                f"v{v_id}",
                status_text,
                str(tr_samples),
                str(a_cnt),
                f"{(a_rate*100):.1f}%",
                f"{a_score:.3f}",
                created
            ])
        if len(model_rows) == 1:
            model_rows.append(["No hay modelos registrados", "-", "-", "-", "-", "-", "-"])

        t_models = Table(model_rows, colWidths=[55, 65, 60, 60, 85, 75, 120])
        t_models.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#334155')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(t_models)
        elements.append(Spacer(1, 10))

        # 3. Comparación Estático vs Adaptativo
        static_m = get_val(comparison, 'static_model')
        adapt_m = get_val(comparison, 'adaptive_model')
        imprv = get_val(comparison, 'improvement')

        if static_m and adapt_m:
            elements.append(Paragraph("3. Comparación: Modelo Estático (M0) vs Adaptativo (Mn)", section_heading))
            
            s_allow = get_val(static_m, 'allow_rate', 0)
            a_allow = get_val(adapt_m, 'allow_rate', 0)
            delta_allow = get_val(imprv, 'allow_rate_delta', 0)

            s_score = get_val(static_m, 'avg_score', 0)
            a_score = get_val(adapt_m, 'avg_score', 0)
            delta_score = get_val(imprv, 'avg_score_delta', 0)

            s_count = get_val(static_m, 'auth_count', 0)
            a_count = get_val(adapt_m, 'auth_count', 0)

            comp_data = [
                ["Métrica", "Modelo Estático (M0)", f"Modelo Adaptativo (v{get_val(adapt_m, 'version_id')})", "Diferencia"],
                ["Tasa de Permitidos", f"{(s_allow*100):.1f}%", f"{(a_allow*100):.1f}%", f"{(delta_allow*100):+.1f}%"],
                ["Score Promedio", f"{s_score:.3f}", f"{a_score:.3f}", f"{delta_score:+.3f}"],
                ["Total Intentos", str(s_count), str(a_count), "-"]
            ]
            t_comp = Table(comp_data, colWidths=[130, 130, 130, 130])
            t_comp.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#ecfdf5')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#065f46')),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#a7f3d0')),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
                ('TOPPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(t_comp)
            elements.append(Spacer(1, 10))

        # 4. Historial de Adaptaciones (últimos 15)
        elements.append(Paragraph("4. Historial de Eventos de Adaptación (Últimos 15)", section_heading))
        adapt_rows = [["Acción", "Modelo Anterior", "Nuevo Modelo", "Razón", "Fecha"]]
        for ev in (timeline or [])[:15]:
            action_clean = str(get_val(ev, 'action', '')).replace('_', ' ').title()
            old_v_id = get_val(ev, 'old_model_version_id')
            new_v_id = get_val(ev, 'new_model_version_id')
            old_v = f"v{old_v_id}" if old_v_id else "—"
            new_v = f"v{new_v_id}" if new_v_id else "—"
            reason = get_val(ev, 'reason', '—') or '—'
            created = str(get_val(ev, 'created_at', '')).replace('T', ' ')[:19]
            adapt_rows.append([action_clean, old_v, new_v, str(reason)[:40], created])
        
        if len(adapt_rows) == 1:
            adapt_rows.append(["Sin eventos registrados", "-", "-", "-", "-"])

        t_adapt = Table(adapt_rows, colWidths=[110, 75, 75, 140, 120])
        t_adapt.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f8fafc')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#475569')),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(t_adapt)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_excel_report(self, db: Session, user_id: int) -> io.BytesIO:
        """Genera un reporte Excel (.xlsx) completo con pestañas estructuradas."""
        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="1E1B4B", end_color="1E1B4B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1E1B4B")
        sub_font = Font(name="Calibri", size=10, italic=True, color="64748B")
        cell_font = Font(name="Calibri", size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        summary = dashboard_service.get_user_summary(db, user_id)
        auth_metrics = dashboard_service.get_auth_metrics(db, user_id)
        models = dashboard_service.get_model_versions(db, user_id)
        timeline = dashboard_service.get_adaptation_timeline(db, user_id)
        comparison = dashboard_service.get_comparison_static_vs_adaptive(db, user_id)

        username = get_val(summary, 'username', f'User_{user_id}')

        # HOJA 1: Resumen & KPIs
        ws1 = wb.active
        ws1.title = "Resumen General"
        ws1.views.sheetView[0].showGridLines = True

        ws1['A1'] = "TECLEOLLAVE-ADAPT — Reporte Biométrico General"
        ws1['A1'].font = title_font
        ws1['A2'] = f"Usuario: {username} (ID: {user_id}) | Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1['A2'].font = sub_font

        ws1.append([])
        headers1 = ["Métrica", "Valor"]
        ws1.append(headers1)
        for col in range(1, 3):
            cell = ws1.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        far_val = get_val(auth_metrics, 'far', 0)
        frr_val = get_val(auth_metrics, 'frr', 0)
        avg_score_val = get_val(auth_metrics, 'avg_score', 0)

        kpis = [
            ("Usuario", username),
            ("Modelo Activo", f"v{get_val(summary, 'active_model_version', 'N/A')}"),
            ("Total Muestras", get_val(summary, 'total_samples', 0)),
            ("Muestras Enrolamiento", get_val(summary, 'enrollment_samples', 0)),
            ("Muestras Autenticación", get_val(summary, 'auth_samples', 0)),
            ("Intentos Autenticación", get_val(auth_metrics, 'total_attempts', 0)),
            ("Intentos Permitidos", get_val(auth_metrics, 'allow_count', 0)),
            ("Intentos Challenge", get_val(auth_metrics, 'challenge_count', 0)),
            ("Intentos Rechazados", get_val(auth_metrics, 'reject_count', 0)),
            ("FAR (False Acceptance Rate)", f"{(far_val*100):.2f}%"),
            ("FRR (False Rejection Rate)", f"{(frr_val*100):.2f}%"),
            ("Score Promedio Biométrico", round(avg_score_val, 4)),
            ("Total Adaptaciones Exitosas", get_val(summary, 'total_adaptations', 0))
        ]

        for k, v in kpis:
            ws1.append([k, v])
            r = ws1.max_row
            ws1.cell(row=r, column=1).font = cell_font
            ws1.cell(row=r, column=1).border = thin_border
            ws1.cell(row=r, column=2).font = cell_font
            ws1.cell(row=r, column=2).border = thin_border

        ws1.column_dimensions['A'].width = 30
        ws1.column_dimensions['B'].width = 25

        # HOJA 2: Historial de Modelos
        ws2 = wb.create_sheet(title="Modelos Biométricos")
        ws2.views.sheetView[0].showGridLines = True
        headers2 = ["Versión", "Estado Activo", "Muestras Entren.", "Intentos Auth", "Tasa Permitidos (%)", "Score Promedio", "Fecha Creación"]
        ws2.append(headers2)
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for m in models or []:
            ws2.append([
                f"v{get_val(m, 'version_id')}",
                "Sí" if get_val(m, 'is_active') else "No",
                get_val(m, 'training_samples', 0),
                get_val(m, 'auth_count', 0),
                round(get_val(m, 'allow_rate', 0) * 100, 2),
                round(get_val(m, 'avg_score', 0), 4),
                str(get_val(m, 'created_at', '')).split('T')[0]
            ])
            r = ws2.max_row
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r, column=c).font = cell_font
                ws2.cell(row=r, column=c).border = thin_border

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # HOJA 3: Timeline de Adaptaciones
        ws3 = wb.create_sheet(title="Eventos Adaptación")
        ws3.views.sheetView[0].showGridLines = True
        headers3 = ["ID Evento", "Acción / Evento", "Modelo Anterior", "Nuevo Modelo", "Razón", "Fecha"]
        ws3.append(headers3)
        for col in range(1, len(headers3) + 1):
            cell = ws3.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for ev in timeline or []:
            old_v_id = get_val(ev, 'old_model_version_id')
            new_v_id = get_val(ev, 'new_model_version_id')
            old_v = f"v{old_v_id}" if old_v_id else "—"
            new_v = f"v{new_v_id}" if new_v_id else "—"
            ws3.append([
                get_val(ev, 'id'),
                str(get_val(ev, 'action', '')).replace('_', ' ').title(),
                old_v,
                new_v,
                get_val(ev, 'reason', '—'),
                str(get_val(ev, 'created_at', '')).replace('T', ' ')[:19]
            ])
            r = ws3.max_row
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=r, column=c).font = cell_font
                ws3.cell(row=r, column=c).border = thin_border

        for col in ws3.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws3.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # HOJA 4: Comparativa M0 vs Mn
        static_m = get_val(comparison, 'static_model')
        adapt_m = get_val(comparison, 'adaptive_model')
        imprv = get_val(comparison, 'improvement')

        if static_m and adapt_m:
            ws4 = wb.create_sheet(title="Estático vs Adaptativo")
            ws4.views.sheetView[0].showGridLines = True
            headers4 = ["Métrica", "Modelo Estático (M0)", f"Modelo Adaptativo (v{get_val(adapt_m, 'version_id')})", "Diferencia"]
            ws4.append(headers4)
            for col in range(1, len(headers4) + 1):
                cell = ws4.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            s_allow = get_val(static_m, 'allow_rate', 0)
            a_allow = get_val(adapt_m, 'allow_rate', 0)
            delta_allow = get_val(imprv, 'allow_rate_delta', 0)

            s_score = get_val(static_m, 'avg_score', 0)
            a_score = get_val(adapt_m, 'avg_score', 0)
            delta_score = get_val(imprv, 'avg_score_delta', 0)

            s_count = get_val(static_m, 'auth_count', 0)
            a_count = get_val(adapt_m, 'auth_count', 0)

            comp_rows = [
                ("Tasa de Permitidos (%)", round(s_allow*100, 2), round(a_allow*100, 2), f"{(delta_allow*100):+.2f}%"),
                ("Score Promedio", round(s_score, 4), round(a_score, 4), f"{delta_score:+.4f}"),
                ("Total Intentos Evaluados", s_count, a_count, "—")
            ]

            for row in comp_rows:
                ws4.append(list(row))
                r = ws4.max_row
                for c in range(1, len(headers4) + 1):
                    ws4.cell(row=r, column=c).font = cell_font
                    ws4.cell(row=r, column=c).border = thin_border

            for col in ws4.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws4.column_dimensions[col_letter].width = max(max_len + 4, 20)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


report_service = ReportService()
